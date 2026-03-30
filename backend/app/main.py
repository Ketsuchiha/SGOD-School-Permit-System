from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional
import uvicorn
from fastapi.responses import Response
import os
import io
import json
import sqlite3
from pathlib import Path
from .services.storage_service import store_permit_file
from .services.backup_service import create_backup, restore_backup, get_backup_info

app = FastAPI(title="DepEd Cabuyao School Permit Registry API")

DB_PATH = Path(__file__).resolve().parents[1] / "data" / "sgod.db"
UPLOADS_DIR = Path(__file__).resolve().parents[1] / "data" / "uploads"
ENV_PATH = Path(__file__).resolve().parents[1] / ".env"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

app.mount("/api/files", StaticFiles(directory=str(UPLOADS_DIR)), name="files")


def load_env_file() -> None:
    if not ENV_PATH.exists():
        return

    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def get_db_connection() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS schools (
                id TEXT PRIMARY KEY,
                payload TEXT NOT NULL,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.commit()


def load_schools_from_db() -> list[dict]:
    with get_db_connection() as conn:
        rows = conn.execute("SELECT payload FROM schools ORDER BY updated_at DESC").fetchall()

    schools: list[dict] = []
    for row in rows:
        try:
            payload = json.loads(row["payload"])
            if isinstance(payload, dict):
                schools.append(payload)
        except Exception:
            continue
    return schools


def save_schools_to_db(schools: list[dict]) -> None:
    with get_db_connection() as conn:
        conn.execute("DELETE FROM schools")
        conn.executemany(
            "INSERT INTO schools (id, payload, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
            [
                (
                    str(school.get("id") or ""),
                    json.dumps(school, ensure_ascii=True),
                )
                for school in schools
                if isinstance(school, dict) and school.get("id")
            ],
        )
        conn.commit()


@app.on_event("startup")
def on_startup() -> None:
    load_env_file()
    init_db()

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify actual frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class GeocodeRequest(BaseModel):
    address: str
    name: Optional[str] = None

class GeocodeResponse(BaseModel):
    lat: float
    lng: float

class ReverseGeocodeRequest(BaseModel):
    lat: float
    lng: float

class ReportPermitLevels(BaseModel):
    kindergarten: bool = False
    elementary: bool = False
    highSchool: bool = False
    seniorHighSchool: bool = False

class ReportSchoolRecord(BaseModel):
    name: str = ""
    address: str = ""
    permitNumber: str = ""
    schoolYear: str = ""
    status: str = ""
    permitLevels: Optional[ReportPermitLevels] = None
    governmentPermits: Optional[list[dict]] = None

class PermitReportRequest(BaseModel):
    schoolYear: Optional[str] = None
    status: Optional[str] = None
    permitLevel: Optional[str] = None
    schools: list[ReportSchoolRecord] = []


class SchoolsBulkRequest(BaseModel):
    schools: list[dict] = []


@app.get("/api/schools")
async def get_schools():
    return {"schools": load_schools_from_db()}


@app.put("/api/schools/bulk")
async def put_schools(payload: SchoolsBulkRequest):
    try:
        save_schools_to_db(payload.schools)
        return {"ok": True, "count": len(payload.schools)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save schools: {str(e)}")


@app.post("/api/uploads/permit")
async def upload_permit_file(request: Request, file: UploadFile = File(...)):
    file_name = (file.filename or "").strip()
    lower_name = file_name.lower()
    if not lower_name.endswith((".pdf", ".jpg", ".jpeg", ".png")):
        raise HTTPException(status_code=400, detail="Invalid file type. Only PDF and images are supported.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File is empty.")

    try:
        result = store_permit_file(
            file_bytes=content,
            filename=file_name,
            content_type=file.content_type or "application/octet-stream",
            uploads_dir=UPLOADS_DIR,
            backend_base_url=str(request.base_url),
        )
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to store permit file: {str(exc)}")
    finally:
        await file.close()

@app.post("/api/ocr/permit")
async def ocr_permit(file: UploadFile = File(...), targetPage: Optional[int] = Form(None)):
    fname = (file.filename or "").lower()
    if not fname.endswith(('.pdf', '.jpg', '.jpeg', '.png')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only PDF and images are supported.")
    content = await file.read()
    # OCR strategy:
    # 1) Extract selectable text from PDF
    # 2) If text is sparse (scanned pages), fallback to image OCR (PyMuPDF + pytesseract) if available
    try:
        from PyPDF2 import PdfReader
        import io, re
        text = ""
        engine = "pdf-text"
        selected_page_numbers: list[int] = []
        page_scores: list[dict] = []

        def score_page(page_text: str) -> int:
            compact = re.sub(r"\s+", " ", page_text).strip().lower()
            if not compact:
                return 0
            score = 0
            if "government permit" in compact or "permit" in compact:
                score += 8
            if "gp-" in compact or "s-" in compact or "gp" in compact:
                score += 6
            if "school year" in compact:
                score += 5
            if "complete address" in compact:
                score += 4
            if "(school)" in compact:
                score += 4
            if "senior high school" in compact:
                score += 5
            if "junior high school" in compact or "high school" in compact:
                score += 4
            if "elementary" in compact:
                score += 3
            if "kindergarten" in compact or " kg " in compact:
                score += 3
            if re.search(r"\bno\.?\s*[a-z]{1,5}-\d{2,6}\b", compact):
                score += 4
            if re.search(r"\b(shs|e|j|k)\s*-\s*\d{3,6}\b", compact):
                score += 5
            if re.search(r"\b(gp|s|p)\s*-\s*[a-z0-9]{2,6}\b", compact):
                score += 5
            if re.search(r"\b20\d{2}-20\d{2}\b", compact) or re.search(r"\b20\d{2}\s*-\s*20\d{2}\b", compact):
                score += 4
            if "deped" in compact or "departme" in compact:
                score += 3
            if "indorsement" in compact or "endorsement" in compact:
                score -= 2
            return score

        def selected_indices(page_count: int, page_texts: list[str]) -> list[int]:
            if page_count <= 0:
                return []

            if targetPage is not None and 1 <= targetPage <= page_count:
                return [targetPage - 1]

            scored = sorted(
                [(score_page(t), idx) for idx, t in enumerate(page_texts)],
                key=lambda x: x[0],
                reverse=True,
            )
            picked = [idx for score, idx in scored if score > 0][:3]
            if picked:
                return picked
            return list(range(min(2, page_count)))

        page_texts: list[str] = []
        if fname.endswith(".pdf"):
            reader = PdfReader(io.BytesIO(content))
            page_texts = [(page.extract_text() or "") for page in reader.pages]
            page_scores = [
                {"page": idx + 1, "score": score_page(page_texts[idx])}
                for idx in range(len(page_texts))
            ]

            # Prioritize likely permit pages first, then append remaining pages.
            priority_idxs = selected_indices(len(page_texts), page_texts)
            selected_page_numbers = [i + 1 for i in priority_idxs]
            remaining_idxs = [i for i in range(len(page_texts)) if i not in set(priority_idxs)]
            ordered_idxs = priority_idxs + remaining_idxs
            text = "\n".join([page_texts[i] for i in ordered_idxs])

            # Scanned PDF fallback: if very little selectable text OR all scores are 0, try image OCR on all pages
            has_meaningful_match = any(re.search(r"\b(SHS|E|J|K)\s*-\s*\d{2,6}\b", page_texts[i], re.IGNORECASE) for i in priority_idxs if i < len(page_texts))
            if len(re.sub(r"\s+", "", text)) < 150 or (not has_meaningful_match and all(s["score"] == 0 for s in page_scores)):
                try:
                    import fitz  # pymupdf
                    import pytesseract
                    from PIL import Image

                    candidates = [
                        os.getenv("TESSERACT_CMD", "").strip(),
                        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                    ]
                    for candidate in candidates:
                        if candidate and os.path.exists(candidate):
                            pytesseract.pytesseract.tesseract_cmd = candidate
                            break

                    ocr_doc = fitz.open(stream=content, filetype="pdf")
                    ocr_chunks: list[str] = []
                    # OCR likely permit pages first for better extraction quality.
                    ocr_order = priority_idxs if priority_idxs else list(range(len(ocr_doc)))
                    ocr_order += [i for i in range(len(ocr_doc)) if i not in set(ocr_order)]
                    for i in ocr_order:
                        page = ocr_doc.load_page(i)
                        pix = page.get_pixmap(dpi=220)
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        ocr_chunks.append(pytesseract.image_to_string(img) or "")
                    ocr_doc.close()
                    fallback_text = "\n".join(ocr_chunks)
                    if len(re.sub(r"\s+", "", fallback_text)) > len(re.sub(r"\s+", "", text)):
                        text = fallback_text
                        engine = "image-ocr"
                except Exception as ocr_error:
                    # If tesseract not available, try alternative: extract images and use basic vision
                    try:
                        import fitz
                        ocr_doc = fitz.open(stream=content, filetype="pdf")
                        pdf_text_parts = []
                        for page_idx in priority_idxs if priority_idxs else range(len(ocr_doc)):
                            page = ocr_doc.load_page(page_idx)
                            # Try to extract text from images in the PDF using basic methods
                            page_text = page.get_text()
                            if page_text.strip():
                                pdf_text_parts.append(page_text)
                        ocr_doc.close()
                        if pdf_text_parts and len(re.sub(r"\s+", "", "\n".join(pdf_text_parts))) > len(re.sub(r"\s+", "", text)):
                            text = "\n".join(pdf_text_parts)
                            engine = "pdf-recovery"
                    except:
                        pass
        else:
            try:
                import pytesseract
                from PIL import Image

                candidates = [
                    os.getenv("TESSERACT_CMD", "").strip(),
                    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                ]
                for candidate in candidates:
                    if candidate and os.path.exists(candidate):
                        pytesseract.pytesseract.tesseract_cmd = candidate
                        break

                image = Image.open(io.BytesIO(content))
                text = pytesseract.image_to_string(image) or ""
                engine = "image-ocr"
            except Exception:
                text = ""
                engine = "none"

        raw_text = text
        normalized = re.sub(r"\s+", " ", raw_text)

        # Basic anchor extraction
        def find(pattern: str) -> str:
            m = re.search(pattern, normalized, re.IGNORECASE)
            return (m.group(1).strip() if m else "")

        def clean_value(value: str) -> str:
            return re.sub(r"\s+", " ", value).strip(" .,-")

        def normalize_permit_code(code: str) -> str:
            compact = re.sub(r"\s+", "", code).upper()
            compact = compact.replace("\u2013", "-").replace("\u2014", "-")

            # Canonical school-level permit codes.
            school_match = re.match(r"^(SHS|E|J|K)-(\d{2,6})$", compact)
            if school_match:
                return f"{school_match.group(1)}-{school_match.group(2)}"

            # Accept broader government permit formats such as GP-ITA, S-067, P-2020A.
            generic_match = re.match(r"^([A-Z]{1,4})-([A-Z0-9]{2,8})$", compact)
            if generic_match:
                return f"{generic_match.group(1)}-{generic_match.group(2)}"

            return ""

        def extract_school_permit_codes(source: str) -> list[str]:
            return [
                f"{prefix.upper()}-{number}"
                for prefix, number in re.findall(r"\b(SHS|E|J|K)\s*[-\u2013\u2014]\s*(\d{2,6})\b", source, re.IGNORECASE)
            ]
        
        def extract_gp_permits(source: str) -> list[str]:
            """Extract Government Permit codes like GP-ITA, S-067, etc."""
            # Match patterns like GP-ITA, GP-123, S-067, S-ITA
            gp_permits = re.findall(r"\b([GSP]{1,2})\s*[-\u2013\u2014]\s*([A-Z0-9]{2,6})\b", source, re.IGNORECASE)
            return [f"{prefix.upper()}-{code.upper()}" for prefix, code in gp_permits]

        def find_first(patterns: list[str], source: str, flags: int = re.IGNORECASE) -> str:
            for pattern in patterns:
                match = re.search(pattern, source, flags)
                if match:
                    return clean_value(match.group(1))
            return ""

        name = find(r"([A-Z][A-Z0-9\s\.,&'\-]{5,})\s*\(\s*School\s*\)")
        if not name:
            name = find(r"(?:Name\s+of\s+School|This\s+is\s+to\s+certify\s+that)[:\s]+([^\n\.,]+)")
        if not name:
            name = find(r"(?:for\s+)?([A-Z][A-Z0-9\s\.,&'\-]{8,})\s+located\s+at")
        if not name:
            name = find_first([
                r"\n\s*([A-Z][A-Z\s\.,&'\-]{6,})\s*\n\s*\(?\s*School\s*\)?",
                r"\b(ST\.?\s+VINCENT\s+COLLEGE\s+OF\s+CABUYAO)\b",
                r"^([A-Z][A-Z0-9\s\.,&'\-]{8,})",  # First line with caps
            ], raw_text, re.IGNORECASE)
        if not name:
            # Try to find any capitalized text that looks like a school name (at least 10 chars, letters/numbers/spaces)
            potential_names = re.findall(r"[A-Z][A-Z0-9 ]{10,}(?:COLLEGE|SCHOOL|ACADEMY|CENTER|INSTITUTE|HIGH|INC|CORP)", normalized, re.IGNORECASE)
            if potential_names:
                name = clean_value(potential_names[0])
        name = clean_value(name)

        address = ""
        line_anchor = re.search(r"\n\s*([^\n]{8,140})\s*\n\s*\(?\s*Complete\s+Address\s*\)?", raw_text, re.IGNORECASE)
        if line_anchor:
            address = clean_value(line_anchor.group(1))
        if not address:
            address = find(r"([A-Z0-9][A-Z0-9\s\.,'\-/]{6,})\s*\(?\s*Complete\s+Address\s*\)?")
        if not address:
            address = find(r"(?:located\s+at|Address:)[:\s]+([^\n]+?)(?:\.|\s+GP\s*No\.|\s+Government\s*Permit|\s+School\s*Year|$)")
        if not address:
            address = find_first([
                r"\n\s*([^\n]{4,120}Cabuyao\s+City)\s*\n\s*\(?\s*Complete\s+Address\s*\)?",
                r"\b([A-Z][A-Za-z\s,.-]{3,120}Cabuyao\s+City)\b",
            ], raw_text, re.IGNORECASE)
        if not address:
            # Try to find any line with Cabuyao City
            cabuyao_lines = re.findall(r"([^.\n]{10,140}Cabuyao\s+(?:City|Laguna))", normalized, re.IGNORECASE)
            if cabuyao_lines:
                address = clean_value(cabuyao_lines[0])
        address = clean_value(address)

        permit_sample_pairs = re.findall(r"\bNo\.?\s*((?:SHS|E|J|K)\s*[-\u2013\u2014]\s*\d{2,6})\s*,?\s*s\.?\s*(20\d{2})", normalized, re.IGNORECASE)
        permit_candidates = [normalize_permit_code(p) for p, _ in permit_sample_pairs]
        inferred_school_years = [f"{y.strip()}-{int(y.strip()) + 1}" for _, y in permit_sample_pairs if y.strip().isdigit()]

        permit_sample_pairs.extend(
            re.findall(r"(?:government\s+permit\s*\([^\)]*\)\s*)?no\.?\s*((?:SHS|E|J|K)\s*[-\u2013\u2014]\s*\d{2,6})\s*,?\s*s\.?\s*(20\d{2})", normalized, re.IGNORECASE)
        )
        permit_candidates.extend([normalize_permit_code(p) for p, _ in permit_sample_pairs])
        inferred_school_years.extend([f"{y.strip()}-{int(y.strip()) + 1}" for _, y in permit_sample_pairs if y.strip().isdigit()])

        labeled_permit_sections = re.findall(
            r"(?:GP\s*No\.?|Government\s*Permit\s*(?:No\.?|Number\.?|#)?|Permit\s*(?:No\.?|Number\.?|#))\s*[:#-]?\s*([A-Z0-9][A-Z0-9\-./\s]{3,})",
            normalized,
            re.IGNORECASE,
        )
        for section in labeled_permit_sections:
            permit_candidates.extend(extract_school_permit_codes(section))

        permit_candidates.extend(extract_school_permit_codes(normalized))

        # Extract Government Permit codes like GP-ITA, S-067, etc.
        permit_candidates.extend(extract_gp_permits(normalized))

        # Additional flexible permit extraction for various formats
        more_permits = re.findall(r"(?:([A-Z]{1,3})\s*-\s*(\d{3,6}))", normalized, re.IGNORECASE)
        for prefix, number in more_permits:
            if prefix.upper() in ['SHS', 'E', 'J', 'K']:
                permit_candidates.append(f"{prefix.upper()}-{number}")
        
        # Extract from common phrases
        common_phrases = re.findall(r"permit(?:ted)?\s+(?:no\.?|number)?\s*(?:is)?\s*([A-Z]+-\d{2,6})", normalized, re.IGNORECASE)
        permit_candidates.extend(common_phrases)
        
        # Extract from lines that look like permit info
        permit_lines = re.findall(r"^[A-Z]{1,3}-\d{3,6}$", normalized, re.MULTILINE)
        permit_candidates.extend(permit_lines)
        
        # Extract permit numbers in barcode-like format (just digits after J/E/SHS/K)
        # Matches: J026608, E00123, SHS0042, etc.
        barcode_matches = re.findall(r"\b([JEKSHS]{1,3})[\s]*0*([0-9]{4,6})\b", normalized, re.IGNORECASE)
        for prefix, number in barcode_matches:
            # Clean up the prefix - remove any "0" characters that might have been included
            clean_prefix = prefix.replace("0", "").upper()[:3]
            if clean_prefix and clean_prefix in ['SHS', 'E', 'J', 'K']:
                numstr = number.lstrip('0') or '0'
                permit_candidates.append(f"{clean_prefix}-{numstr}")
        
        # Also try to match even without spaces between prefix and number
        # E.g., "J026608" as a continuous string
        tight_barcode = re.findall(r"\b([JE])(\d{6})\b", normalized, re.IGNORECASE)
        for prefix, number in tight_barcode:
            permit_candidates.append(f"{prefix.upper()}-{number.lstrip('0') or number}")
        
        tight_barcode_shs = re.findall(r"\b(SHS|K)(\d{6})\b", normalized, re.IGNORECASE)
        for prefix, number in tight_barcode_shs:
            permit_candidates.append(f"{prefix.upper()}-{number.lstrip('0') or number}")

        deduped_permits = []
        seen = set()
        for p in permit_candidates:
            normalized_permit = normalize_permit_code(p)
            key = normalized_permit.strip().upper()
            if key and key not in seen:
                deduped_permits.append(normalized_permit.strip())
                seen.add(key)

        school_year_matches = re.findall(r"(?:effective\s+this\s+school\s+year\s+)?(20\d{2}\s*[-/]\s*20\d{2})", normalized, re.IGNORECASE)
        school_year_matches.extend(
            re.findall(r"effective\s+this\s+school\s+year\s+([0-9]{4}\s*[-/]\s*[0-9]{4})", normalized, re.IGNORECASE)
        )
        # Additional school year patterns
        school_year_matches.extend(re.findall(r"school\s+year\s+([0-9]{4}\s*[-/]\s*[0-9]{4})", normalized, re.IGNORECASE))
        school_year_matches.extend(re.findall(r"year\s+([0-9]{4}\s*[-/]\s*[0-9]{4})", normalized, re.IGNORECASE))
        school_year_matches.extend(re.findall(r"s\.?\s*([0-9]{4}\s*[-/]\s*[0-9]{4})", normalized, re.IGNORECASE))
        school_year_matches.extend(re.findall(r"\(([0-9]{4}\s*[-/]\s*[0-9]{4})\)", normalized, re.IGNORECASE))
        # Match patterns like "2017 -" or "2017-" (incomplete year ranges)
        school_year_matches.extend(re.findall(r"\b(20\d{2})\s*[-/]\s*$", normalized, re.IGNORECASE | re.MULTILINE))
        school_year_matches.extend(re.findall(r"(\d{4})\s*[-/]\s*(\d{4})", normalized, re.IGNORECASE))
        school_year_matches.extend(re.findall(r"effective\s+this\s+school\s+year\s+(\d{4})\s*-\s*(\d{4})", normalized, re.IGNORECASE))
        
        # Infer school years from single years found in document
        single_years = re.findall(r"\b(20\d{2})\b", normalized)
        for year in single_years:
            if year.isdigit():
                year_int = int(year)
                inferred_school_years.append(f"{year_int}-{year_int + 1}")
        
        school_years = []
        year_seen = set()
        for y in (school_year_matches + inferred_school_years):
            # Handle tuple results from certain regexes
            if isinstance(y, tuple):
                y = f"{y[0]}-{y[1]}"
            clean = re.sub(r"\s+", "", y).replace("/", "-")
            # Remove trailing dash if incomplete
            clean = re.sub(r"-$", "", clean)
            if clean and (re.match(r"^\d{4}-\d{4}$", clean) or re.match(r"^\d{4}$", clean)):
                # If just a single year, expand it
                if re.match(r"^\d{4}$", clean):
                    year_int = int(clean)
                    clean = f"{year_int}-{year_int + 1}"
                if clean not in year_seen:
                    school_years.append(clean)
                    year_seen.add(clean)

        def contains_any(source: str, patterns: list[str]) -> bool:
            return any(re.search(p, source, re.IGNORECASE) for p in patterns)

        default_levels = {
            "kindergarten": contains_any(normalized, [r"\bkindergarten\b", r"\bK\b\s*-\s*kindergarten", r"kg", r"\(K\)"]),
            "elementary": contains_any(normalized, [r"\belementary\b", r"\bE\b\s*-\s*elementary", r"elem", r"\(E\)"]),
            "highSchool": contains_any(normalized, [r"junior\s+high\s+school", r"\bJHS\b", r"\bJ\b\s*-", r"junior high", r"\(J\)"]),
            "seniorHighSchool": contains_any(normalized, [r"senior\s+high\s+school", r"\bSHS\b", r"senior high school program", r"senior high", r"SHS program", r"\(SHS\)"]),
        }

        if default_levels["seniorHighSchool"]:
            default_levels["highSchool"] = False
        
        # If we found SHS in permit codes, mark as senior high school
        if any(code.startswith("SHS") for code in deduped_permits):
            default_levels["seniorHighSchool"] = True
            default_levels["highSchool"] = False
        # If we found E in permit codes, mark as elementary
        if any(code.startswith("E-") for code in deduped_permits):
            default_levels["elementary"] = True
        # If we found J in permit codes, mark as junior high school
        if any(code.startswith("J-") for code in deduped_permits):
            default_levels["highSchool"] = True
        # If we found K in permit codes, mark as kindergarten
        if any(code.startswith("K-") for code in deduped_permits):
            default_levels["kindergarten"] = True

        strand_patterns = {
            "STEM": [r"\bSTEM\b", r"science\s*,?\s*technology\s*,?\s*engineering\s*(?:and|&)\s*mathematics"],
            "ABM": [r"\bABM\b", r"accountancy\s*,?\s*business\s*(?:and|&)\s*management"],
            "HUMSS": [r"\bHUMSS\b", r"humanities\s*(?:and|&)\s*social\s+sciences"],
            "GAS": [r"\bGAS\b", r"general\s+academic\s+strand"],
            "TVL-ICT": [r"\bTVL[-\s]*ICT\b", r"information\s+and\s+communications?\s+technology"],
            "TVL-HE": [r"\bTVL[-\s]*HE\b", r"home\s+economics", r"\bhome\s+ec\b"],
            "TVL-IA": [r"\bTVL[-\s]*IA\b", r"industrial\s+arts"],
            "ARTS-DESIGN": [r"\bARTS[-\s]*DESIGN\b", r"arts\s*(?:and|&)\s*design"],
            "SPORTS": [r"\bSPORTS\b", r"sports\s+track"],
            "TVL": [r"\bTVL\b", r"technical\s+(?:and\s+)?vocational(?:\s+livelihood)?"],
        }
        detected_strands = [
            strand
            for strand, patterns in strand_patterns.items()
            if any(re.search(pattern, normalized, re.IGNORECASE) for pattern in patterns)
        ]
        if detected_strands:
            default_levels["seniorHighSchool"] = True

        permits = []
        if deduped_permits:
            for idx, permit_no in enumerate(deduped_permits):
                permits.append(
                    {
                        "permitNumber": permit_no,
                        "schoolYear": school_years[idx] if idx < len(school_years) else (school_years[0] if school_years else ""),
                        "permitLevels": default_levels,
                        "shsStrands": detected_strands,
                    }
                )
        else:
            permits.append(
                {
                    "permitNumber": "",
                    "schoolYear": school_years[0] if school_years else "",
                    "permitLevels": default_levels,
                    "shsStrands": detected_strands,
                }
            )

        primary = permits[0] if permits else {
            "permitNumber": "",
            "schoolYear": "",
            "permitLevels": {
                "kindergarten": False,
                "elementary": False,
                "highSchool": False,
                "seniorHighSchool": False,
            },
            "shsStrands": [],
        }

        school_year = find(r"(\d{4}\s*-\s*\d{4})")
        selected_scores = [
            item["score"] for item in page_scores if item["page"] in set(selected_page_numbers)
        ]
        if not selected_scores and page_scores:
            selected_scores = [item["score"] for item in page_scores[:3]]
        
        # Boost confidence based on how many fields we successfully extracted
        fields_found = 0
        total_critical_fields = 5
        if name: fields_found += 1
        if address: fields_found += 1
        if primary["permitNumber"]: fields_found += 1
        resolved_school_year = primary["schoolYear"] or school_year
        if resolved_school_year: fields_found += 1
        if any(primary["permitLevels"].values()): fields_found += 1
        
        # Calculate confidence with a minimum floor
        # Base it more on field extraction than page scoring since page scoring is unreliable for scanned PDFs
        fields_confidence = fields_found / total_critical_fields
        
        if selected_scores:
            base_confidence = min(1.0, max(0.0, sum(selected_scores) / (len(selected_scores) * 12.0)))
        else:
            base_confidence = 0.1 if engine == "image-ocr" else 0.0  # Give some credit to image OCR
        
        # Weight: 30% on page scoring, 70% on field extraction success
        # This makes it more robust for scanned documents
        confidence = min(1.0, max(0.1 * (fields_found > 0), (base_confidence * 0.3) + (fields_confidence * 0.7)))

        missing_fields = []
        if not name:
            missing_fields.append("name")
        if not address:
            missing_fields.append("address")
        if not primary["permitNumber"]:
            missing_fields.append("permitNumber")
        resolved_school_year = primary["schoolYear"] or school_year
        if not resolved_school_year:
            missing_fields.append("schoolYear")
        if not any(primary["permitLevels"].values()):
            missing_fields.append("permitLevels")

        return {
            "name": name,
            "address": address,
            "permitNumber": primary["permitNumber"],
            "schoolYear": resolved_school_year,
            "permitLevels": primary["permitLevels"],
            "shsStrands": primary["shsStrands"],
            "permits": permits,
            "ocrEngine": engine,
            "ocrDiagnostics": {
                "selectedPages": selected_page_numbers,
                "confidence": confidence,
                "missingFields": missing_fields,
                "totalPages": len(page_texts) if fname.endswith('.pdf') else 1,
                "topPageScores": sorted(page_scores, key=lambda item: item["score"], reverse=True)[:5],
            },
        }
    except Exception as e:
        # Fallback: return empty fields but keep endpoint responsive
        return {
            "name": "",
            "address": "",
            "permitNumber": "",
            "schoolYear": "",
            "permitLevels": {
                "kindergarten": False,
                "elementary": False,
                "highSchool": False,
                "seniorHighSchool": False,
            },
            "shsStrands": [],
            "permits": [],
            "ocrEngine": "none",
            "ocrDiagnostics": {
                "selectedPages": [],
                "confidence": 0.0,
                "missingFields": ["name", "address", "permitNumber", "schoolYear", "permitLevels"],
                "totalPages": 0,
                "topPageScores": [],
            },
        }

@app.post("/api/geocode", response_model=GeocodeResponse)
async def geocode(request: GeocodeRequest):
    try:
        from importlib import import_module
        import re
        geocoders = import_module("geopy.geocoders")
        Nominatim = geocoders.Nominatim
        geolocator = Nominatim(user_agent="deped-cabuyao-school-registry")

        address = (request.address or "").strip()
        name = (request.name or "").strip()

        # Accept direct coordinate input, e.g. "14.2722, 121.1239".
        coord_match = re.search(r"(-?\d{1,3}(?:\.\d+)?)\s*[,\s]\s*(-?\d{1,3}(?:\.\d+)?)", address)
        if coord_match:
            lat = float(coord_match.group(1))
            lng = float(coord_match.group(2))
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                return {"lat": lat, "lng": lng}

        normalized_address = re.sub(r"\s+", " ", address).strip()
        queries = []
        if name and address:
            queries.append(f"{name}, {normalized_address}, Cabuyao City, Laguna, Philippines")
            queries.append(f"{name}, {normalized_address}, Cabuyao, Laguna, Philippines")
            queries.append(f"{name}, {normalized_address}, Laguna, Philippines")
        if address:
            queries.append(f"{normalized_address}, Cabuyao City, Laguna, Philippines")
            queries.append(f"{normalized_address}, Cabuyao, Laguna, Philippines")
            queries.append(f"{normalized_address}, Laguna, Philippines")
            queries.append(normalized_address)
        if name:
            queries.append(f"{name}, Cabuyao City, Laguna, Philippines")
            queries.append(f"{name}, Laguna, Philippines")

        deduped_queries = []
        seen = set()
        for q in queries:
            key = q.lower().strip()
            if key and key not in seen:
                deduped_queries.append(q)
                seen.add(key)

        def score_location(query_text: str, found_address: str) -> int:
            score = 0
            found = (found_address or "").lower()
            query = (query_text or "").lower()

            if "cabuyao" in found:
                score += 10
            if "laguna" in found:
                score += 8
            if "philippines" in found:
                score += 3

            if name and name.lower() in found:
                score += 6

            tokens = [token for token in re.split(r"[^a-z0-9]+", query) if len(token) >= 4]
            overlap = sum(1 for token in tokens if token in found)
            score += min(overlap, 8)

            return score

        best_location = None
        best_score = -1
        laguna_viewbox = [(121.00, 14.42), (121.24, 14.17)]

        for query in deduped_queries:
            try:
                location = geolocator.geocode(
                    query,
                    timeout=12,
                    exactly_one=True,
                    country_codes="ph",
                    viewbox=laguna_viewbox,
                    bounded=False,
                    addressdetails=True,
                )
            except Exception:
                location = None

            if not location:
                continue

            found_address = getattr(location, "address", "") or ""
            found_lat = float(location.latitude)
            found_lng = float(location.longitude)
            if not (4 <= found_lat <= 22 and 116 <= found_lng <= 127):
                continue

            score = score_location(query, found_address)
            if score > best_score:
                best_score = score
                best_location = location

            if "cabuyao" in found_address.lower() and score >= 18:
                break

        if not best_location:
            raise HTTPException(status_code=404, detail="Address not found")
        return {"lat": best_location.latitude, "lng": best_location.longitude}
    except ImportError:
        raise HTTPException(status_code=503, detail="Geocoding service unavailable")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Geocoding failed: {str(e)}")


@app.post("/api/reverse-geocode")
async def reverse_geocode(request: ReverseGeocodeRequest):
    try:
        from importlib import import_module
        geocoders = import_module("geopy.geocoders")
        Nominatim = geocoders.Nominatim
        geolocator = Nominatim(user_agent="deped-cabuyao-school-registry")

        lat = float(request.lat)
        lng = float(request.lng)

        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            raise HTTPException(status_code=400, detail="Invalid coordinates")

        location = geolocator.reverse(
            f"{lat}, {lng}",
            timeout=12,
            language="en",
            addressdetails=True,
            zoom=18,
        )

        if not location:
            raise HTTPException(status_code=404, detail="Address not found")

        return {
            "address": getattr(location, "address", "") or "",
            "lat": lat,
            "lng": lng,
        }
    except ImportError:
        raise HTTPException(status_code=503, detail="Geocoding service unavailable")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Reverse geocoding failed: {str(e)}")

@app.get("/api/reports/permits")
async def get_permit_report(
    schoolYear: Optional[str] = Query(None),
    status: Optional[str] = Query(None)
):
    # Minimal CSV export to avoid heavy XLSX dependencies
    csv = "School Name,District,Status,School Year,Permit No\n"
    # No DB yet; return empty dataset with headers only
    return Response(
        content=csv.encode("utf-8"),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=school_permits.csv"}
    )


@app.post("/api/reports/permits")
async def create_permit_report(payload: PermitReportRequest):
    try:
        from openpyxl import Workbook
        from collections import defaultdict

        def normalize_levels(levels: Optional[dict]) -> dict:
            source = levels or {}
            return {
                "kindergarten": bool(source.get("kindergarten")),
                "elementary": bool(source.get("elementary")),
                "highSchool": bool(source.get("highSchool")),
                "seniorHighSchool": bool(source.get("seniorHighSchool")),
            }

        def level_text(levels: dict) -> str:
            if not levels:
                return ""
            tags = []
            if levels.get("kindergarten"):
                tags.append("Kindergarten")
            if levels.get("elementary"):
                tags.append("Elementary")
            if levels.get("highSchool"):
                tags.append("Junior High School")
            if levels.get("seniorHighSchool"):
                tags.append("Senior High School")
            return ", ".join(tags)

        def permit_matches_level(levels: dict, selected_level: str) -> bool:
            if not selected_level or selected_level == "all":
                return True
            return bool(levels.get(selected_level, False))

        wb = Workbook()
        # Remove the default empty sheet created by openpyxl
        wb.remove(wb.active)

        selected_year = (payload.schoolYear or "").strip()
        selected_status = (payload.status or "").strip().lower()
        selected_level = (payload.permitLevel or "all").strip()
        headers = ["School Name", "School Address", "Government Permit", "School Year", "Year Level", "Status"]

        def get_permit_number(school: ReportSchoolRecord) -> str:
            num = (school.permitNumber or "").strip()
            if not num and school.governmentPermits:
                for permit in school.governmentPermits:
                    candidate = str(permit.get("permitNumber", "")).strip()
                    if candidate:
                        return candidate
            return num

        def auto_size(ws_obj):
            for col in ws_obj.columns:
                max_len = 0
                for cell in col:
                    cell_val = "" if cell.value is None else str(cell.value)
                    if len(cell_val) > max_len:
                        max_len = len(cell_val)
                ws_obj.column_dimensions[col[0].column_letter].width = min(max(14, max_len + 2), 60)

        # Filter by status first
        filtered = []
        for school in payload.schools:
            school_status = (school.status or "").strip().lower()
            if selected_status and selected_status != "all" and school_status != selected_status:
                continue
            filtered.append(school)

        report_rows: list[dict] = []
        for school in filtered:
            permit_history = school.governmentPermits or []
            permit_levels_obj = school.permitLevels
            if permit_levels_obj is None:
                permit_levels_data = None
            elif hasattr(permit_levels_obj, "model_dump"):
                permit_levels_data = permit_levels_obj.model_dump()
            elif hasattr(permit_levels_obj, "dict"):
                permit_levels_data = permit_levels_obj.dict()
            else:
                permit_levels_data = None

            fallback_levels = normalize_levels(permit_levels_data)
            fallback_permit = {
                "permitNumber": (school.permitNumber or "").strip(),
                "schoolYear": (school.schoolYear or "").strip(),
                "permitLevels": fallback_levels,
            }

            has_fallback_data = bool(
                fallback_permit["permitNumber"]
                or fallback_permit["schoolYear"]
                or any(fallback_levels.values())
            )

            merged_permits = []
            for permit in permit_history:
                if not isinstance(permit, dict):
                    continue

                raw_levels = permit.get("permitLevels", {})
                normalized = {
                    "permitNumber": str(permit.get("permitNumber", "")).strip(),
                    "schoolYear": str(permit.get("schoolYear", "")).strip(),
                    "permitLevels": normalize_levels(raw_levels if isinstance(raw_levels, dict) else {}),
                }
                merged_permits.append(normalized)

            fallback_duplicate = any(
                (p.get("permitNumber", "").strip().lower() == fallback_permit["permitNumber"].lower())
                and (p.get("schoolYear", "").strip() == fallback_permit["schoolYear"])
                for p in merged_permits
            )

            if has_fallback_data and not fallback_duplicate:
                merged_permits.append(fallback_permit)

            if not merged_permits:
                merged_permits.append(fallback_permit)

            for permit in merged_permits:
                permit_year = (permit.get("schoolYear", "") or "").strip()
                permit_levels = permit.get("permitLevels", {})

                if selected_year and permit_year != selected_year:
                    continue
                if not permit_matches_level(permit_levels, selected_level):
                    continue

                report_rows.append({
                    "school_name": (school.name or "").strip(),
                    "school_address": (school.address or "").strip(),
                    "permit_number": (permit.get("permitNumber") or "").strip() or get_permit_number(school),
                    "school_year": permit_year,
                    "year_level": level_text(permit_levels),
                    "status": (school.status or "").strip(),
                })

        if selected_year:
            # Single worksheet for a specific school year
            ws = wb.create_sheet(title=selected_year[:31])
            ws.append(headers)
            for row in report_rows:
                ws.append([
                    row["school_name"],
                    row["school_address"],
                    row["permit_number"],
                    row["school_year"],
                    row["year_level"],
                    row["status"],
                ])
            auto_size(ws)
        else:
            # One worksheet per school year, sorted; "No Year" sheet for blanks
            year_buckets: dict = defaultdict(list)
            for row in report_rows:
                yr = row["school_year"] or "No Year"
                year_buckets[yr].append(row)

            for yr in sorted(year_buckets.keys()):
                sheet_title = yr[:31]
                ws = wb.create_sheet(title=sheet_title)
                ws.append(headers)
                for row in year_buckets[yr]:
                    ws.append([
                        row["school_name"],
                        row["school_address"],
                        row["permit_number"],
                        row["school_year"],
                        row["year_level"],
                        row["status"],
                    ])
                auto_size(ws)

        # Ensure at least one sheet exists
        if not wb.worksheets:
            ws = wb.create_sheet(title="No Data")
            ws.append(headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename_year = selected_year if selected_year else "all-years"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=school-permit-report-{filename_year}.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")


@app.get("/api/backup")
async def download_backup():
    """
    Create and download a backup of all schools and permit files.
    Returns a zip file containing schools.json and all uploads.
    """
    try:
        schools = load_schools_from_db()
        backup_data = create_backup(schools, UPLOADS_DIR)
        return Response(
            content=backup_data,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=sgod-backup.zip"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup creation failed: {str(e)}")


@app.post("/api/backup/restore")
async def restore_from_backup(file: UploadFile = File(...)):
    """
    Restore schools and files from a backup zip file.
    New files are restored; existing files are skipped to avoid overwrites.
    """
    try:
        backup_content = await file.read()
        if not backup_content:
            raise HTTPException(status_code=400, detail="Backup file is empty")
        
        restore_result = restore_backup(backup_content, UPLOADS_DIR)
        
        # Save restored schools to database
        if restore_result["schools"]:
            save_schools_to_db(restore_result["schools"])
        
        return {
            "ok": True,
            "schools_restored": len(restore_result["schools"]),
            "files_restored": restore_result["files_restored"],
            "files_skipped": restore_result["files_skipped"],
            "timestamp": restore_result["timestamp"],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup restore failed: {str(e)}")
    finally:
        await file.close()


@app.post("/api/backup/info")
async def get_backup_details(file: UploadFile = File(...)):
    """
    Get information about a backup file without restoring it.
    """
    try:
        backup_content = await file.read()
        if not backup_content:
            raise HTTPException(status_code=400, detail="Backup file is empty")
        
        info = get_backup_info(backup_content)
        return info
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read backup info: {str(e)}")
    finally:
        await file.close()


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
